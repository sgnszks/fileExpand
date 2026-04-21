"""
API 路由测试

覆盖:
- 上传接口参数校验
- 文件类型过滤
- 正常处理流程
- 下载接口
- 支持格式查询
"""
import io
import os

import pytest
import openpyxl


class TestUploadAPI:
    """上传 API 测试。"""

    def test_no_file(self, client):
        """缺少文件参数应返回 400。"""
        resp = client.post('/api/upload', data={'multiplier': '2'})
        assert resp.status_code == 400
        data = resp.get_json()
        assert data['success'] is False

    def test_no_multiplier(self, client):
        """缺少倍数参数应返回 400。"""
        buf = io.BytesIO(b'dummy')
        resp = client.post('/api/upload', data={
            'file': (buf, 'test.xlsx'),
        }, content_type='multipart/form-data')
        assert resp.status_code == 400

    def test_unsupported_extension(self, client):
        """不支持的文件类型应返回 400。"""
        buf = io.BytesIO(b'dummy')
        resp = client.post('/api/upload', data={
            'file': (buf, 'test.txt'),
            'multiplier': '2',
        }, content_type='multipart/form-data')
        assert resp.status_code == 400
        data = resp.get_json()
        assert '不支持' in data['error']

    def test_valid_xlsx_upload(self, client):
        """有效的 XLSX 上传应返回成功。"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Test'
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post('/api/upload', data={
            'file': (buf, 'test.xlsx'),
            'multiplier': '2',
        }, content_type='multipart/form-data')

        data = resp.get_json()
        assert data['success'] is True
        assert data['data']['actual_multiplier'] > 1.0

    def test_invalid_multiplier(self, client):
        """非法倍数应返回 400。"""
        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        resp = client.post('/api/upload', data={
            'file': (buf, 'test.xlsx'),
            'multiplier': '0.5',
        }, content_type='multipart/form-data')

        assert resp.status_code == 400


class TestSupportedFormatsAPI:
    """支持格式查询 API 测试。"""

    def test_get_formats(self, client):
        """应返回支持的文件格式列表。"""
        resp = client.get('/api/supported-formats')
        assert resp.status_code == 200
        data = resp.get_json()
        assert 'formats' in data
        extensions = [f['extension'] for f in data['formats']]
        assert '.xlsx' in extensions
        assert '.pdf' in extensions


class TestDownloadAPI:
    """下载 API 测试。"""

    def test_nonexistent_file(self, client):
        """请求不存在的文件应返回 404。"""
        resp = client.get('/api/download/nonexistent_file.xlsx')
        assert resp.status_code == 404

    def test_path_traversal(self, client):
        """路径穿越尝试应被阻止。"""
        resp = client.get('/api/download/..%2F..%2Fetc%2Fpasswd')
        assert resp.status_code in (403, 404)
